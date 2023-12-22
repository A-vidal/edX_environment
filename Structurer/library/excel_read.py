# from control import *
import os
import re
import openpyxl

# the excel pages to read
MC_sections = ["Template", "Tests"]

# finds out if the value of a cell is a function or a reference
def is_func(x: str) -> bool:
    if type(x) == str:
        return x.startswith("=") or x.startswith("=+")
    else:
        return False

# finds the excel file in the MC directory
def find_excel(MC_path: str, MC_code: str) -> str:
    # MC_path: path of the course
    # MC_code: code of the course
    return [i for i in os.listdir(MC_path) if re.match(f".*{MC_code}.*xlsx$", i)][0]

class Excel_read:
    # constructor: read the excel with openpyxl
    def __init__(self, filename: str, template: str, tests: str) -> None:
        # filename: path to the excel file
        # template: sheet name of the course structure
        # tests: sheet name of the course tests
        self.filename = filename
        self.WB = openpyxl.load_workbook(filename= filename, read_only= False, keep_vba= True, keep_links= True)
        self.MC_Tests = self.WB.get_sheet_by_name(tests)
        self.MC_Template = self.WB.get_sheet_by_name(template)
        try:
            self.MC_code = self.MC_Template.cell(1,3).value.strip(" ")
        except:
            self.MC_code = input("Please introduce the MC_code: ").strip(" ")
    
    # return the global information of the course
    def get_info(self) -> tuple[str, str, str]:
        MC_title = self.MC_Template.cell(2,3).value
        MC_lecturer = self.MC_Template.cell(3,3).value
        MC_inst = self.MC_Template.cell(4,3).value
        return (MC_title, MC_lecturer, MC_inst)

    # return a structure that represents the course test
    def get_tests(self) -> list[list]:
        retorn = []
        t_class = [int, str, int, str, int, str, int, str, float, str, str, str, str, str, int, str, str, str, str, str, str, str, str, str, str]
        #           A    B    C    D    E    F    g    H     I     J    K    L    M    N    O    P    Q    R    S    T    U    V    W    X    Y 
        cabecera = True
        k = 0
        for i in self.MC_Tests.iter_rows():
            if cabecera:
                cabecera = False
                continue

            if all([x.value == None for x in i]):
                continue
            
            k += 1
            if k == 0:
                break
            t_values = []
            
            for j in i:
                
                if i.index(j) < 6:
                    if j.internal_value == None:
                        t_values.append(None)

                    elif is_func(j.internal_value):
                        
                        value = j.internal_value

                        while is_func(value):
                            if value.find("!") != -1:
                                c_range = value.split("!")[-1]
                                value = self.MC_Template[c_range.replace("$","")].value
                            else:
                                c_range = value.lstrip("=").lstrip("+")
                                value = self.MC_Tests[c_range.replace("$","")].value
                        
                        if type(value) == t_class[i.index(j)]:
                            t_values.append(value)
                        
                        elif type(value) == str and t_class[i.index(j)] == int:
                            try:
                                t_values.append(int(re.findall("[0-9]+", value)[0]))
                            except:
                                print("test error:",value)
                                print(j.internal_value)
                                t_values.append(-1)
                        else:
                            t_values.append(t_class[i.index(j)](value))

                    elif type(j.value) == t_class[i.index(j)]:
                        t_values.append(j.value)

                    else:
                        raise f"Can't read the ({k}, {i.index(j)}) value"
                else:
                    t_values.append(j.value)
                continue

            if any([t != None for t in t_values]):
                retorn.append(t_values)
            
        for i in retorn:
            for j in i[12:]:
                if type(j) == str:
                    j = j[0].upper() + j[1:]
        return retorn

    # converts the content of the test cell into a number or None
    def clean_test(self, x) -> float | int | None:
        # x: cell content
        t = type(x)
        if x == None:
            return None
        
        if t == float or t == int:
            return x
        
        if is_func(x):
            x = str(x)
            fnd = re.findall("[A-Z][0-9]+", x)
            print("clean test:", fnd)
            for i in fnd:
                value = self.MC_Template[i].value
                while is_func(value):
                    value = self.MC_Template[value.strip("=").strip("=+").strip("=-")].value
                ind = x.index(i)
                x = x[:ind] + "{0:f}".format(value) + x[ind+len(i):]
                # x.replace(i, "{0:f}".format(value))
            return eval(x.lstrip("=+").lstrip("="))
        
        if t == str:
            find = re.findall("[0-9]+.[0-9]*", x)
            if len(find) == 0:
                return None
            
            return float(find[0])

    # return a structure that represents the course structure
    def get_units(self) -> list[list]:
    
        sec_n, sub_n = (1,) * 2

        sec, sub = ("",) * 2

        Units = []

        # parse a individual unit
        def unit(row: list):
            # row: a list of cells where the unit is contained
            video = row[3].value
            if type(row[3].value) == str:
                if len(row[3].value) != 11:
                    # Create a structure to upload the video
                    atr = ["","", -1]
                    atr[0] = f"./Section {sec_n}/{self.MC_code}-S{sec_n}-Sub{sub_n}-U{row[1].value}.mp4"
                    atr[1] = ".".join([str(i) for i in [self.MC_code, sec_n, sub_n, row[1].value]]) + " - " + row[2].value
                    atr[2] = row[3].row
                    video = atr
            try:
                test = self.clean_test(row[4].value)
            except IndexError:
                test = -1.0
                print(f"Error in test in row: {(row[2].value, row[4].value, row[5].value)}\nFind it in S{sec_n}-Sub{sub_n}-U{row[1].value}")
        
            Units.append([sec_n, sec, sub_n, sub, row[1].value, row[2].value, video, test, row[5].value])

        
        for row in self.MC_Template.iter_rows():
            
            header = str(row[0].value)
            
            if bool(re.match("^Section.*[0-9]+.*$",header)):
                sec_n = int(re.findall("[0-9]+", header)[0])
                sec = row[1].value
                sub = row[1].value
                sub_n = 1
                continue
            elif bool(re.match("^Subsection.*$",header)):
                sub_n = int(row[1].value)
                sub = row[2].value
                continue
            elif bool(re.match("^.*Unit.*$",header)):
                unit(row)
        
        return Units
    
    # upload the videos with the YT_bot
    def upload_videos(self, bot, units: list[list], desc_plantilla: str) -> bool:
        # bot: the YT_bot (need to be available)
        # units: the units structure of the course
        # desc_plantilla: the video description with a "{0}" where put the unit code
        print(f"Description: \n{desc_plantilla}\n")
        def upload_video(file: str, title: str, row: int) -> str:
            try:
                v_code = ((file.split("/")[-1]).split(".")[0])[5:]
                video = bot.subir_video(file, title, desc_plantilla.format(v_code))
                print(title, "-", video)
                self.MC_Template.cell(row, 4).value = video
                assert self.MC_Template.cell(row, 4).value == video, "El video no se ha gruadado en el Excel"
                return video
            except:
                print("Err: upload_video",(file, title, row))
                self.save()
                self.close()
                assert False, "ERROR: problema al subir los videos"
    
        print("Uploading Videos...:")
        for i in units:
            if type(i[6]) == list:
                i[6] = upload_video(*i[6])
        
        self.save()
        self.close()

    # save the excel file
    def save(self, filename: str = "") -> None:
        # filename: were the save is realised
        if filename == "":
            self.WB.save(self.filename)
        else:
            self.WB.save(filename)
    
    # close the excel file
    def close(self) -> None:
        self.WB.close()
    
    # in case the object is destroyed
    def __delattr__(self, __name: str) -> None:
        try:
            self.save()
            self.close()
        except:
            pass
        del self.WB
