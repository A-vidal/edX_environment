from library.control import PDFtoHTML
import openpyxl

DATOS_GENERALES = [
        [
            'MC_code',
            'MC_title', #MC_name
            'datetime.datetime(2024, 1, 1, 0, 0)',
            'datetime.datetime(2024, 1, 1, 0, 0)',
            "English" # ingles?
        ],
        [
            'MC_lecturer', # Intro_dic["Lecturers"],
            None,
            'Description_dic["Background of the proposed micro-credential"]' + "\n\n" + 'Description_dic["Overview of the micro-credential"]',
            'Description_dic["Prerequisites"]',
            'Description_dic["Learning objectives"]',
        ],
        [
            'MC_inst', # Intro_dic["Institution"],
            'Description_dic["ECTS"]'.strip() + " ECTS",
            None,
            None,
            'Evaluacion'
        ],
        [
            None,
            "Z-Posgrade",
            "2024-001",
            "No",
            'logo'
        ],
        [
            'datetime.datetime(2024, 1, 1, 0, 0)',
            None,
            None,
            None,
            "Si"
        ]
    ]

# create dicctionary with a unique name for every type of test
def dict_tareas(Units : list[list]) -> dict[tuple, tuple]:
    # Units: the units structure of the excel_read
    # output: a diccionary with the name an ocurences os a type of test
    tareas = [(u[-1], u[-2]) for u in Units if u[-2] != None]

    tipos = dict([(i, (0,)) for i in set(tareas)])
    nombres = [i[0] for i in tipos]
    dic_nombres = dict([(i, nombres.count(i)) for i in set(nombres)])

    for i in tipos.keys():
        if dic_nombres[i[0]] > 1:
            ext = f" ({dic_nombres[i[0]]})"
            dic_nombres[i[0]] -= 1
        else:
            ext = ""
        tipos[i] = (i[0] + ext, tareas.count(i))
    return tipos

# resize a table of a sheet as wanted
def ajustar_tabla(Sheet: openpyxl.worksheet, table_key: str, max_row: int):
    # Sheet: WorkBookSheet which has the table inside
    # table_key: the name of the table
    # max_row: starting from the 1 row, which row is the lastest of the table
    table = Sheet.tables[table_key]
    rang = openpyxl.worksheet.cell_range.CellRange(table.ref)
    rang.max_row = max_row
    table.ref = rang.coord
    return rang.coord

class Excel_write:
    # constructor: write a new excel with openpyxl
    def __init__(self, MC_path: str) -> None:
        # MC_code: the code of the course
        self.MC_code = MC_path.split("/")[2]
        self.filename = MC_path.split("/")[0]+"/Excel_output/"+self.MC_code+".xlsm"

        WB_out = openpyxl.load_workbook(filename= MC_path.split("/")[0]+"/Data/TemplateXNF.xlsm", read_only= False, keep_vba= True, keep_links= True)
        WB_out.save(self.filename)
        WB_out.close()

        self.WB = openpyxl.load_workbook(filename= self.filename, read_only= False, keep_vba= True, keep_links= True)
        self.DatosGenerales = self.WB.get_sheet_by_name("DatosGenerales")
        self.Tipo_de_tarea = self.WB.get_sheet_by_name("TipodeTarea")
        self.Unidades = self.WB.get_sheet_by_name("Unidades")
        self.Leccion = self.WB.get_sheet_by_name("Leccion")
        self.Problemas = self.WB.get_sheet_by_name("Problemas")
    
    # write the general information of the course
    def upload_datos(self, estructura: list[list]) -> None:
        # estructura: a DATOS_GENERALES like structure
        for j in range(5): 
            for i in range(len(estructura)):
                try:
                    self.DatosGenerales.cell((1+i)*2, (1+j)).value = estructura[i][j]
                except ValueError:
                    print(i, j)
    
    # write the types of problems there are going to be
    def upload_tareas(self, tareas: dict[tuple, tuple]) -> None:
        # tareas: a dict_tareas return like dictionary
        i = 2
        for t in tareas.keys():
            i += 1
            self.Tipo_de_tarea.cell(i, 1).value = tareas[t][0]
            self.Tipo_de_tarea.cell(i, 2).value = tareas[t][0]
            self.Tipo_de_tarea.cell(i, 3).value = t[1]
            self.Tipo_de_tarea.cell(i, 4).value = 0
            self.Tipo_de_tarea.cell(i, 5).value = 1
            self.Tipo_de_tarea.cell(i, 6).value = "Terminada"
            self.Tipo_de_tarea.cell(i, 7).value = 1
            self.Tipo_de_tarea.cell(i, 8).value = "Por Estudiante"
            self.Tipo_de_tarea.cell(i, 9).value = tareas[t][1]
            
        ajustar_tabla(self.Tipo_de_tarea, list(self.Tipo_de_tarea.tables.keys())[0], i)
        self.tipos_de_tarea = tareas
    
    # write the principal units there are going to be
    def upload_units(self, Units: list[list]):
        # Units: the units structure of the excel_read
        i = 2
        # sec_unit = set([(u[0], u[2]) for u in Units])
        last = Units[0]
        for unit in Units:
            
            if unit[:4] != last[:4]:
                i += 1
            
            tarea = 0
            if unit[-2] != None:
                tarea = self.tipos_de_tarea[(unit[-1], unit[-2])][0]
            
            self.Unidades.cell(i, 2).value = unit[0]
            self.Unidades.cell(i, 3).value = unit[1]
            self.Unidades.cell(i, 4).value = unit[2]
            self.Unidades.cell(i, 5).value = unit[3]
            self.Unidades.cell(i, 6).value = tarea
            self.Unidades.cell(i, 7).value = None
            self.Unidades.cell(i, 8).value = None
            last = unit
        
        ajustar_tabla(self.Unidades, "Unidades", i)
    
    # write all units there are going to be
    def upload_leccion(self, Units: list[list]):
        # Units: the units structure of the excel_read
        i = 1
        for u in Units:
            i += 1
            self.Leccion.cell(i, 1).value = u[0]
            self.Leccion.cell(i, 2).value = f"{u[0]}:{u[1]}"
            self.Leccion.cell(i, 3).value = u[2]
            self.Leccion.cell(i, 4).value = f"{u[0]}:{u[2]}:{u[3]}"
            self.Leccion.cell(i, 5).value = u[4]
            self.Leccion.cell(i, 6).value = u[5]
            self.Leccion.cell(i, 7).value = None
            self.Leccion.cell(i, 8).value = None
            self.Leccion.cell(i, 9).value = PDFtoHTML(u[0], u[-1], u[5], f"{self.MC_code}-S{u[0]}-Sub{u[2]}-U{u[4]}") if u[-1] != None and u[-2] == None and u[6] != None else None
            self.Leccion.cell(i,10).value = u[6]
            self.Leccion.cell(i,11).value = None
            self.Leccion.cell(i,12).value = PDFtoHTML(u[0], u[-1], u[5], f"{self.MC_code}-S{u[0]}-Sub{u[2]}-U{u[4]}") if u[-1] != None and u[-2] == None and u[6] == None else None
            self.Leccion.cell(i,13).value = None if u[-2] != None else 1
            self.Leccion.cell(i,14).value = ""
            self.Leccion.cell(i,15).value = f"{u[0]}:{u[2]}:{u[4]}:{u[5]}"
        
        ajustar_tabla(self.Leccion, "Leccion", i)

    # write all tests there are going to be
    def upload_problema(self, Tests: list[list]):
        # Tests: the test structure of the excel_read
        i = 1
        for t in Tests:
            i += 1
            self.Problemas.cell(i, 1).value = t[0]
            self.Problemas.cell(i, 2).value = f"{t[0]}:{t[1]}"
            self.Problemas.cell(i, 3).value = t[2]
            self.Problemas.cell(i, 4).value = f"{t[0]}:{t[2]}:{t[3]}"
            self.Problemas.cell(i, 5).value = t[4]
            self.Problemas.cell(i, 6).value = f"{t[0]}:{t[2]}:{t[4]}:{t[5]}"
            self.Problemas.cell(i, 7).value = None
            self.Problemas.cell(i, 8).value = "Terminada"
            self.Problemas.cell(i, 9).value = 1
            self.Problemas.cell(i,10).value = None
            self.Problemas.cell(i,11).value = None
            self.Problemas.cell(i,12).value = "MultiChoice"
            self.Problemas.cell(i,13).value = t[12]
            self.Problemas.cell(i,14).value = None
            self.Problemas.cell(i,15).value = t[14]
            for j in range(15, len(t)):
                self.Problemas.cell(i, (1+j)).value = t[j]
        
        ajustar_tabla(self.Problemas, "Problemas", i)

    # uploads all the excel in one call
    def upload(self, datos_generales: list[list], Units: list[list], Tests: list[list]):
        # datos_generales: a DATOS_GENERALES like structure
        # Units: the units structure of the excel_read
        # Tests: the test structure of the excel_read
        self.upload_datos(datos_generales)
        self.upload_tareas(dict_tareas(Units))
        self.upload_units(Units)
        self.upload_leccion(Units)
        self.upload_problema(Tests)

    # save the excel file
    def save(self, filename: str = "") -> None:
        # filename: were the save is realised
        if filename == "":
            self.WB.save(self.filename)
        else:
            self.WB.save(filename)
    
    # close the excel file
    def close(self) -> None:
        self.WB.close()
    
    # in case the object is destroyed
    def __delattr__(self, __name: str) -> None:
        try:
            self.save()
            self.close()
        except:
            pass
        del self.WB